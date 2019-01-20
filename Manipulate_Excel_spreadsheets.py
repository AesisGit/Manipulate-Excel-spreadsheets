import openpyxl

theFile = openpyxl.load_workbook('Customers1.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGHIJKL":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == "telephone":
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name

def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    print(letter)
    return letter


def get_all_values_by_cell_letter(letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            #take old data and send it to fixing
            telephoneNo = fix_telephone_format(currentSheet[cell_name].value)
            #put new data in cell


            #print(letter + "1")
            if cell_name == (letter + "1"):
                #print(letter + "0")
                #print("aaaaa")
                currentSheet[cell_name].value = "telephone"
            else:
                currentSheet[cell_name].value = telephoneNo



            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))


#Remove first + from the telephone number
def remove_plus_from_tel(telephoneNo):
    if telephoneNo[0] == "+":
        telephoneNo = telephoneNo[1:len(telephoneNo)]
    return telephoneNo

#Remove Swedish country code
def remove_country_code(telephoneNo):
    if telephoneNo[0:2] == "46":
        telephoneNo = telephoneNo[2:len(telephoneNo)]
    elif telephoneNo[0:3] == "046":
        telephoneNo = telephoneNo[3:len(telephoneNo)]
    elif telephoneNo[0:4] == "0046":
        telephoneNo = telephoneNo[4:len(telephoneNo)]
    return telephoneNo

#IF 0 is missing at beggining place it
def place_zero_at_first(telephoneNo):
    if telephoneNo[0] != "0":
        telephoneNo = "0" + telephoneNo
    return telephoneNo

def fix_telephone_format(telephoneNo):
    telephoneNo =remove_plus_from_tel(telephoneNo)
    telephoneNo = remove_country_code(telephoneNo)
    telephoneNo = place_zero_at_first(telephoneNo)
    #telephoneNo = remove _all_characters(telephoneNo)
    return telephoneNo




for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())
    letter = get_column_letter(specificCellLetter)


    get_all_values_by_cell_letter(letter)

    theFile.save("Customers2.xlsx")


